"""
Template 7: Texas Instruments — Material Declaration (.xlsx)
Sheet 'Detail', rows with Substance + CAS Number + Amount (mg).
Skips Sub-Total, Total, and section header rows.
"""
import csv
from decimal import Decimal
import openpyxl

SKIP_SUBSTANCES = {"Sub-Total", "Total", None}


def detect(file_path: str) -> bool:
    """Return True if this xlsx matches the Texas Instruments material declaration layout."""
    try:
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if "Detail" not in wb.sheetnames:
            return False
        ws = wb["Detail"]
        for row in ws.iter_rows(max_row=5, values_only=True):
            if any("Texas Instruments" in str(c) for c in row if c):
                return True
    except Exception:
        return False
    return False


def extract(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb["Detail"]

    results = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        # Locate the header row
        if not header_found:
            if row[0] == "Component" and row[1] == "Substance":
                header_found = True
            continue

        component, substance, cas, amount_mg, *_ = row

        # Stop at footer
        if component == "Important Note":
            break

        # Skip section headers (component filled, substance empty) and totals
        if substance in SKIP_SUBSTANCES or not cas or amount_mg is None:
            continue

        results.append({
            "material":  str(substance).strip(),
            "substance": str(cas).strip(),
            "weight_mg": f"{Decimal(str(amount_mg)):.10f}".rstrip("0").rstrip("."),
        })

    return results


def save_csv(records: list[dict], output_path: str) -> None:
    if not records:
        return
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["material", "substance", "weight_mg"])
        writer.writeheader()
        writer.writerows(records)
