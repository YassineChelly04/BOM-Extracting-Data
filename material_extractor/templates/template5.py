"""
Template 5: Nexperia — Product Content Declaration (.xlsx)
Sheet 'Declaration', rows with a CAS number and numeric Mass (mg).
Skips Total/subtotal rows.
"""
import csv
from decimal import Decimal
import openpyxl

SKIP_MATERIALS = {"Adhesive Total", "Die Total", "Lead Frame Total",
                  "Mould Compound Total", "Post-Plating Total", "Wire Total", "Total"}


def detect(file_path: str) -> bool:
    """Return True if this xlsx matches the Nexperia Product Content Declaration layout."""
    try:
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return False
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if "Declaration" not in wb.sheetnames:
            return False
        ws = wb["Declaration"]
        for row in ws.iter_rows(max_row=5, values_only=True):
            if any("Nexperia" in str(c) for c in row if c):
                return True
    except Exception:
        return False
    return False


def extract(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb["Declaration"]

    results = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        # Locate the header row
        if not header_found:
            if row[0] == "Material" and row[3] == "Substance":
                header_found = True
            continue

        material, _, _, substance, cas, mass_mg, *_ = row

        # Stop at footer rows
        if material in (None, "Info", "Disclaimer") and substance is None:
            if material in ("Info", "Disclaimer"):
                break
            continue

        # Skip subtotal/total rows and rows without a substance
        if material in SKIP_MATERIALS or not substance or not cas or mass_mg is None:
            continue

        results.append({
            "material":  str(substance).strip(),
            "substance": str(cas).strip(),
            "weight_mg": f"{Decimal(str(mass_mg)):.10f}".rstrip("0").rstrip("."),
        })

    return results


def save_csv(records: list[dict], output_path: str) -> None:
    if not records:
        return
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["material", "substance", "weight_mg"])
        writer.writeheader()
        writer.writerows(records)
