"""Extractor factory and template-specific extractors."""
from __future__ import annotations

from pathlib import Path
from typing import Any
from decimal import Decimal

import pdfplumber
import openpyxl

from material_extractor.core.extractor import (
    PDFTableExtractor, ExcelSheetExtractor, BaseExtractor
)
from material_extractor.models import (
    TemplateDefinition, ExtractionResult, MaterialRecord, SourceType
)


# --- Registry Pattern (replaces if/elif chain) ---

EXTRACTOR_REGISTRY: dict[str, type[BaseExtractor]] = {}

def register_extractor(name: str):
    """Decorator to register an extractor class."""
    def decorator(cls):
        EXTRACTOR_REGISTRY[name] = cls
        return cls
    return decorator


def create_extractor(template: TemplateDefinition) -> BaseExtractor:
    """Create appropriate extractor for template using registry."""
    extractor_type = template.extraction.get("extractor_type", "pdf_table")
    extractor_class = template.extraction.get("extractor_class", "")

    if extractor_class in EXTRACTOR_REGISTRY:
        return EXTRACTOR_REGISTRY[extractor_class](template)

    if extractor_type == "excel":
        return GenericExcelExtractor(template)

    return GenericPDFTableExtractor(template)


# --- Template-Specific Extractors ---

@register_extractor("MolexTemplateExtractor")
class MolexTemplateExtractor(PDFTableExtractor):
    """Template 1: Molex/TE Connectivity BOM Annex 3 table with case sizes."""

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        if len(table) < 3:
            return []

        records = []
        case_sizes = [c for c in table[1][4:] if c]
        skip_rows = {"Homogeneous Level Weight", "Total Weight"}

        for row_idx, row in enumerate(table[3:], start=3):
            if not row or len(row) < 5:
                continue
            level, material, substance, cas, *weights = row

            if not material or (level and level in skip_rows):
                continue

            weight_dict = dict(zip(case_sizes, weights))
            target_size = self.config.get("target_size", "0402")
            weight_str = weight_dict.get(target_size, "0")

            records.append(self._create_record(
                material=material.strip(),
                substance=substance.strip() if substance else "",
                weight_mg=weight_str,
                case_size=target_size,
                raw_material=material,
                raw_substance=substance or "",
                raw_weight=weight_str,
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("WurthTemplateExtractor")
class WurthTemplateExtractor(PDFTableExtractor):
    """Template 2: Würth Elektronik — Semi-Component breakdown + part table."""

    def extract(self, file_path: Path) -> ExtractionResult:
        result = self._create_result()
        result.source_file = str(file_path)

        try:
            with pdfplumber.open(file_path) as pdf:
                pages = self.metadata.page_indices or [0]
                for page_idx in pages:
                    if page_idx >= len(pdf.pages):
                        continue
                    page = pdf.pages[page_idx]
                    tables = page.extract_tables()
                    result.tables_found += len(tables)
                    result.pages_processed += 1

                    if len(tables) < 3:
                        continue

                    part_table = tables[2]
                    raw_mass = str(part_table[1][1] if len(part_table) > 1 and len(part_table[1]) > 1 else "0")
                    total_mass_g = float(raw_mass.lower().replace("g", "").replace("mg", "").strip())

                    material_table = tables[1]
                    for row_idx, row in enumerate(material_table[1:], start=1):
                        if len(row) < 5:
                            continue
                        _, _, substance, _, avg_mass, _ = row
                        if not substance or not avg_mass:
                            continue

                        weight_mg = round(float(avg_mass) / 100 * total_mass_g * 1000, 4)

                        result.records.append(self._create_record(
                            material=substance.replace("\n", " ").strip(),
                            substance="",
                            weight_mg=weight_mg,
                            raw_material=substance,
                            raw_substance="",
                            raw_weight=str(weight_mg),
                            page_number=page_idx,
                            row_index=row_idx
                        ))

                    result.tables_processed += 1

            result.success = len(result.records) > 0
        except Exception as e:
            result.success = False
            result.errors.append(str(e))

        return result


@register_extractor("MolexComplianceExtractor")
class MolexComplianceExtractor(PDFTableExtractor):
    """Template 3: Molex Product Compliance Declaration — substance rows in grams."""

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []

        for row_idx, row in enumerate(table[1:], start=1):
            if len(row) < 5:
                continue
            name, type_, cas, _, mass_g = row
            if type_ != "Substance" or not mass_g:
                continue

            weight_mg = float(mass_g) * 1000

            records.append(self._create_record(
                material=name.strip(),
                substance=cas.strip() if cas else "",
                weight_mg=weight_mg,
                raw_material=name,
                raw_substance=cas or "",
                raw_weight=str(weight_mg),
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("VishayTemplateExtractor")
class VishayTemplateExtractor(PDFTableExtractor):
    """Template 9: Vishay Material Declaration Sheet."""

    SKIP_SUBSTANCES = {"", None, "Miscellaneous", "miscellaneous"}
    SKIP_CAS = {"System", "", None}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        for row_idx, row in enumerate(table, start=0):
            if len(row) < 6:
                continue
            substance = row[3]
            cas = row[4]
            weight = row[5]

            if not substance or not cas or not weight:
                continue
            if cas in self.SKIP_CAS or substance in self.SKIP_SUBSTANCES:
                continue
            try:
                float(weight)
            except (ValueError, TypeError):
                continue

            records.append(self._create_record(
                material=substance.strip(),
                substance=cas.strip(),
                weight_mg=weight,
                raw_material=substance,
                raw_substance=cas,
                raw_weight=weight,
                page_number=page,
                row_index=row_idx
            ))
        return records


@register_extractor("YageoTemplateExtractor")
class YageoTemplateExtractor(PDFTableExtractor):
    """Template 13: YAGEO Material Composition Declaration — Content(mg) in col 6."""

    SKIP_PARTS = {"Total", "", None}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        for row_idx, row in enumerate(table[1:], start=1):
            if len(row) < 7:
                continue
            no, part, part_wt, material, substance, cas, content_mg, *_ = row

            if part in self.SKIP_PARTS and not substance:
                continue
            if not substance or not content_mg:
                continue

            cas_fixed = str(cas or "").strip()
            if cas_fixed.startswith(")"):
                cas_fixed = cas_fixed[1:]

            try:
                float(content_mg)
            except (ValueError, TypeError):
                continue

            records.append(self._create_record(
                material=substance.replace("\n", " ").strip(),
                substance=cas_fixed,
                weight_mg=content_mg.strip(),
                raw_material=substance,
                raw_substance=cas_fixed,
                raw_weight=content_mg,
                page_number=page,
                row_index=row_idx
            ))
        return records


@register_extractor("NexperiaExcelExtractor")
class NexperiaExcelExtractor(ExcelSheetExtractor):
    """Template 5: Nexperia Product Content Declaration (.xlsx)."""

    SKIP_MATERIALS = {"Adhesive Total", "Die Total", "Lead Frame Total",
                      "Mould Compound Total", "Post-Plating Total", "Wire Total", "Total"}

    def extract(self, file_path: Path) -> ExtractionResult:
        result = self._create_result()
        result.source_file = str(file_path)

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if "Declaration" not in wb.sheetnames:
                result.errors.append("Sheet 'Declaration' not found")
                return result
            ws = wb["Declaration"]

            records = []
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    if row[0] == "Material" and row[3] == "Substance":
                        header_found = True
                    continue

                material, _, _, substance, cas, mass_mg, *_ = row

                if material in (None, "Info", "Disclaimer") and substance is None:
                    if material in ("Info", "Disclaimer"):
                        break
                    continue

                if material in self.SKIP_MATERIALS or not substance or not cas or mass_mg is None:
                    continue

                records.append(self._create_record(
                    material=str(substance).strip(),
                    substance=str(cas).strip(),
                    weight_mg=Decimal(str(mass_mg)),
                    raw_material=str(substance),
                    raw_substance=str(cas),
                    raw_weight=str(mass_mg),
                ))

            result.records = records
            result.success = len(records) > 0
            result.tables_processed = 1
        except Exception as e:
            result.errors.append(str(e))

        return result


@register_extractor("TIExcelExtractor")
class TIExcelExtractor(ExcelSheetExtractor):
    """Template 7: Texas Instruments Material Declaration (.xlsx — Detail sheet)."""

    SKIP_SUBSTANCES = {"Sub-Total", "Total", None}

    def extract(self, file_path: Path) -> ExtractionResult:
        result = self._create_result()
        result.source_file = str(file_path)

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if "Detail" not in wb.sheetnames:
                result.errors.append("Sheet 'Detail' not found")
                return result
            ws = wb["Detail"]

            records = []
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    if row[0] == "Component" and row[1] == "Substance":
                        header_found = True
                    continue

                component, substance, cas, amount_mg, *_ = row

                if component == "Important Note":
                    break

                if substance in self.SKIP_SUBSTANCES or not cas or amount_mg is None:
                    continue

                records.append(self._create_record(
                    material=str(substance).strip(),
                    substance=str(cas).strip(),
                    weight_mg=Decimal(str(amount_mg)),
                    raw_material=str(substance),
                    raw_substance=str(cas),
                    raw_weight=str(amount_mg),
                ))

            result.records = records
            result.success = len(records) > 0
            result.tables_processed = 1
        except Exception as e:
            result.errors.append(str(e))

        return result


@register_extractor("IHSFormatExtractor")
class IHSFormatExtractor(PDFTableExtractor):
    """Template 6: IHS Normalized Format — sub_part / substance / weight."""

    SKIP_SUBSTANCES = {"Total Weight\n(mg)", "Total Weight"}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        if table_idx != 1:
            return records

        for row_idx, row in enumerate(table[4:], start=4):
            if len(row) < 6:
                continue
            sub_part, _, _, substance, cas, weight_mg, *_ = row

            if not substance or not weight_mg:
                continue
            if substance in self.SKIP_SUBSTANCES or "cid" in str(substance):
                continue
            try:
                float(weight_mg)
            except (ValueError, TypeError):
                continue

            records.append(self._create_record(
                material=substance.replace("\n", " ").strip(),
                substance=cas.strip() if cas else "",
                weight_mg=weight_mg,
                raw_material=substance,
                raw_substance=cas or "",
                raw_weight=str(weight_mg),
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("LiteonTemplateExtractor")
class LiteonTemplateExtractor(PDFTableExtractor):
    """Template 8: LITEON Material Composition Declaration."""

    SKIP_ELEMENTS = {"ttl wt = 1.60 mg", None, ""}
    SKIP_CAS = {"Trade secret", ""}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        for row_idx, row in enumerate(table[3:], start=3):
            if len(row) < 8:
                continue
            _, _, _, _, element, cas, _, weight_mg = row[:8]

            if not element or element in self.SKIP_ELEMENTS:
                continue
            if not cas or cas.strip() in self.SKIP_CAS:
                continue
            if not weight_mg:
                continue

            try:
                val = Decimal(str(weight_mg).strip())
            except Exception:
                continue

            records.append(self._create_record(
                material=element.strip(),
                substance=cas.strip(),
                weight_mg=val,
                raw_material=element,
                raw_substance=cas,
                raw_weight=str(weight_mg),
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("HarwinTemplateExtractor")
class HarwinTemplateExtractor(PDFTableExtractor):
    """Template 10: Harwin Material Declaration — multiple variants, weight in grams."""

    SKIP_SUBSTANCES = {"Other Impurities", "", None}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        if not table or table[0][0] != "Homogeneous Material Location":
            return records

        for row_idx, row in enumerate(table[1:], start=1):
            if len(row) < 5:
                continue
            location, weight_g, tolerance, substance, cas = row[:5]

            if not substance or substance in self.SKIP_SUBSTANCES or not cas:
                continue
            try:
                w = Decimal(str(weight_g).strip())
                if w <= 0:
                    continue
            except Exception:
                continue

            weight_mg = w * 1000

            records.append(self._create_record(
                material=substance.strip(),
                substance=cas.strip(),
                weight_mg=weight_mg,
                raw_material=substance,
                raw_substance=cas,
                raw_weight=str(weight_mg),
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("EDSExtractor")
class EDSExtractor(PDFTableExtractor):
    """Template 11: EDS Ingredient Composition Table (Murata-style)."""

    SKIP_PARTS = {"Total （Design value）", None, ""}

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        for row_idx, row in enumerate(table[3:], start=3):
            if len(row) < 7:
                continue
            part, _, _, _, substance, cas, mass_mg, *_ = row

            if part in self.SKIP_PARTS and not substance:
                continue
            if not substance or not cas or not mass_mg:
                continue
            try:
                float(mass_mg)
            except (ValueError, TypeError):
                continue

            records.append(self._create_record(
                material=substance.strip(),
                substance=cas.strip(),
                weight_mg=mass_mg.strip(),
                raw_material=substance,
                raw_substance=cas,
                raw_weight=mass_mg,
                page_number=page,
                row_index=row_idx
            ))

        return records


@register_extractor("StackpoleTemplateExtractor")
class StackpoleTemplateExtractor(PDFTableExtractor):
    """Template 12: Stackpole SEI — multi-value cells (newline-separated)."""

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        for row_idx, row in enumerate(table[1:], start=1):
            if len(row) < 4:
                continue
            bom_item, materials, cas_numbers, weights, *_ = row

            if not materials or not weights or "Total Weight" in str(bom_item or ""):
                continue

            mat_list = str(materials).split("\n")
            cas_list = str(cas_numbers).split("\n") if cas_numbers else []
            wgt_list = str(weights).split("\n")

            for i, mat in enumerate(mat_list):
                mat = mat.strip()
                cas = cas_list[i].strip() if i < len(cas_list) else ""
                wgt = wgt_list[i].strip() if i < len(wgt_list) else ""

                if not mat or not wgt:
                    continue
                try:
                    float(wgt)
                except (ValueError, TypeError):
                    continue

                records.append(self._create_record(
                    material=mat,
                    substance=cas,
                    weight_mg=wgt,
                    raw_material=mat,
                    raw_substance=cas,
                    raw_weight=wgt,
                    page_number=page,
                    row_index=row_idx
                ))

        return records


@register_extractor("GenericPDFTableExtractor")
class GenericPDFTableExtractor(PDFTableExtractor):
    """Generic fallback — auto-detect material/substance/weight columns."""

    def _process_table(self, table: list[list[str]], page: int, table_idx: int) -> list[MaterialRecord]:
        records = []
        header = [str(c).lower() for c in table[0] if c]

        mat_col = self._find_column(header, ["material", "component", "name"])
        sub_col = self._find_column(header, ["cas", "substance", "chemical", "formula"])
        wt_col = self._find_column(header, ["weight", "mass", "mg", "g", "amount", "%"])

        if mat_col is None:
            return records

        for row_idx, row in enumerate(table[1:], start=1):
            if not row or len(row) <= mat_col:
                continue
            material = row[mat_col] if row[mat_col] else ""
            if not material.strip():
                continue

            substance = row[sub_col] if sub_col is not None and len(row) > sub_col else ""
            weight = row[wt_col] if wt_col is not None and len(row) > wt_col else "0"

            records.append(self._create_record(
                material=material.strip(),
                substance=substance.strip() if substance else "",
                weight_mg=weight,
                raw_material=material,
                raw_substance=substance or "",
                raw_weight=weight,
                page_number=page,
                row_index=row_idx
            ))

        return records

    def _find_column(self, headers: list[str], keywords: list[str]) -> int | None:
        for i, h in enumerate(headers):
            for kw in keywords:
                if kw in h:
                    return i
        return None


class GenericExcelExtractor(ExcelSheetExtractor):
    """Generic Excel extractor — auto-detect columns by name."""

    def _process_dataframe(self, df) -> list[MaterialRecord]:
        records = []
        for row_idx, row in df.iterrows():
            material = str(row.get("Material", row.get("material", ""))).strip()
            if not material:
                continue

            records.append(self._create_record(
                material=material,
                substance=str(row.get("CAS", row.get("Substance", ""))).strip(),
                weight_mg=str(row.get("Weight", row.get("Weight_mg", "0"))),
                raw_material=material,
                raw_substance=str(row.get("CAS", row.get("Substance", ""))),
                raw_weight=str(row.get("Weight", row.get("Weight_mg", "0"))),
                row_index=row_idx
            ))
        return records